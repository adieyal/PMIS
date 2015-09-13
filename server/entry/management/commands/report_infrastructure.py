import iso8601
import json

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from libs.database.database import UUID
from libs.database.backend import connection
from libs.database.database import Project
from openpyxl import load_workbook, Workbook
from reports.views import normalize_district, normalize, _currency, _safe_float, _safe_int, _expenditure_for_month_v2, _progress_for_month_v2

sign = lambda x: x and (1, -1)[x<0]

clusters = {
    'E': 'Education',
    'H': 'Health',
    'SD': 'Social Development',
    'CSR': 'Culture, Sports and Recreation',
    'CSSL': 'Community Safety, Security and Liaison',
    'EDET': 'Economic Development, Environment and Tourism',
}

def calculate_rows(cluster, year, phase):
    # Input is financial year, so it's for the period ending March of the next year
    year += 1
    month = 3

    def actual_progress(p):
        if p.phase == 'implementation':
            return _safe_float(_progress_for_month_v2(p.actual, year, month)) * 100 or 0
        else:
            return None

    def planned_progress(p):
        if p.phase == 'implementation':
            return _safe_float(_progress_for_month_v2(p.planning, year, month)) * 100 or 0
        else:
            return None

    def calculate_progress(p):
        actual = actual_progress(project)
        planned = planned_progress(project)

        if actual is not None and planned is not None:
            return sign(actual - planned)

    columns = {
        'B': 'contract',
        'C': 'description',
        'D': 'scope',
        'E': lambda project: normalize_district(project.district) or 'Unknown',
        'F': lambda project: normalize(project.municipality) or 'Unknown',
        'G': 'location',
        'H': lambda project: _safe_float(project.total_anticipated_cost),
        'J': 'source',
        'P': 'expenditure_to_date',
        'Q': 'total_previous_expenses',
        'R': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year-1, 4)),
        'S': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year-1, 5)),
        'T': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year-1, 6)),
        'U': lambda project: (_safe_float(_expenditure_for_month_v2(project.actual, year-1, 4) +
                                        _expenditure_for_month_v2(project.actual, year-1, 5) +
                                        _expenditure_for_month_v2(project.actual, year-1, 6))),
        'V': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year-1, 7)),
        'W': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year-1, 8)),
        'X': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year-1, 9)),
        'Y': lambda project: (_safe_float(_expenditure_for_month_v2(project.actual, year-1, 7) +
                                        _expenditure_for_month_v2(project.actual, year-1, 8) +
                                        _expenditure_for_month_v2(project.actual, year-1, 9))),
        'Z': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year-1, 10)),
        'AA': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year-1, 11)),
        'AB': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year-1, 12)),
        'AC': lambda project: (_safe_float(_expenditure_for_month_v2(project.actual, year-1, 10) +
                                        _expenditure_for_month_v2(project.actual, year-1, 11) +
                                        _expenditure_for_month_v2(project.actual, year-1, 12))),
        'AD': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year, 1)),
        'AE': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year, 2)),
        'AF': lambda project: _safe_float(_expenditure_for_month_v2(project.actual, year, 3)),
        'AG': lambda project: (_safe_float(_expenditure_for_month_v2(project.actual, year, 1) +
                                        _expenditure_for_month_v2(project.actual, year, 2) +
                                        _expenditure_for_month_v2(project.actual, year, 3))),
        'AM': 'planned_start',
        'AN': 'actual_start',
        'AO': 'planned_completion',
        'AP': 'revised_completion',
        'AQ': 'actual_completion',
        'AT': lambda project: actual_progress(project),
        'AU': lambda project: calculate_progress(project),
        'AV': 'comments',
        'AW': 'principal_agent',
        'AX': 'contractor',
        'AY': 'manager',
    }

    project_ids = Project.list()

    rows = []

    for project_id in project_ids:
        project = Project.get(project_id)

        if project.cluster == 'Department of %s' % cluster and project.phase == phase:
            row = {}

            for column, calc in columns.iteritems():
                if type(calc) == str:
                    value = getattr(project, calc)
                else:
                    value = calc(project)

                row[column] = value

            rows.append(row)

    return rows

class Command(BaseCommand):
    # args = 'filename'
    help = 'Create Infrastructure Report'

    def translate_cluster(self, cluster):
        return cluster.lower().replace('department of ', '').replace(' ', '-').replace(',', '')

    def add_arguments(self, parser):
        parser.add_argument('year', nargs='+', type=int)

    def handle(self, *args, **kwargs):
        year = int(args[0])

        workbook = load_workbook('fixtures/infrastructure-template.xlsx')

        for code, title in clusters.iteritems():
            sheet = workbook.create_sheet(title=code)

            count = 0
            for row in workbook['Sheet1'].iter_rows():
                count += 1

                if count < 8:
                    sheet.append([ i for i in row ])
                else:
                    break

            sheet['B4'] = '%s Infrastructure Projects - April %s' % (title, year)

            sheet['R6'] = 'Apr-%s' % year
            sheet['S6'] = 'May-%s' % year
            sheet['T6'] = 'Jun-%s' % year

            sheet['V6'] = 'Jul-%s' % year
            sheet['W6'] = 'Aug-%s' % year
            sheet['X6'] = 'Sep-%s' % year

            sheet['Z6'] = 'Oct-%s' % year
            sheet['AA6'] = 'Nov-%s' % year
            sheet['AB6'] = 'Dec-%s' % year

            sheet['AD6'] = 'Jan-%s' % (year + 1)
            sheet['AE6'] = 'Feb-%s' % (year + 1)
            sheet['AF6'] = 'Mar-%s' % (year + 1)

            sheet['AH6'] = 'Exp. To date\n(%s/%s)\nR\'000' % (year, year + 1)
            sheet['AJ6'] = 'Total Projected Exp\n(%s/%s)\nR\'000' % (year, year + 1)
            sheet['AK6'] = 'Balance of Budget\n(%s/%s)\nR\'000' % (year, year + 1)

            rows = calculate_rows(title, year, 'planning')

            if len(rows):
                sheet.append(['PLANNING'])

                count = 0

                for row in rows:
                    count += 1
                    row['A'] = count
                    sheet.append(row)


            rows = calculate_rows(title, year, 'implementation')

            if len(rows):
                sheet.append(['IMPLEMENTATION'])

                count = 0

                for row in rows:
                    count += 1
                    row['A'] = count
                    sheet.append(row)

        workbook.remove_sheet(workbook['Sheet1'])
        workbook.remove_sheet(workbook['Sheet2'])
        workbook.remove_sheet(workbook['Sheet3'])

        workbook.save(filename='data/infrastructure.xlsx')
